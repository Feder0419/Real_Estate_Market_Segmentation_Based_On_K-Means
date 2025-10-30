from coordTransform_utils import bd09_to_wgs84
import openpyxl

path=r"副本sfxiaoqu.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook["sfxiaoqu"]

for i in range(2,9614):
	x=sheet.cell(row=i,column=14).value
	y=sheet.cell(row=i,column=15).value
	coordinates=bd09_to_wgs84(float(x),float(y))
	sheet["Q"+str(i)]=coordinates[0]
	sheet["R"+str(i)]=coordinates[1]

workbook.save(path)